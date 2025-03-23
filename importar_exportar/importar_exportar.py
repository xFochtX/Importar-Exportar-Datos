import pandas as pd
import pickle
from pathlib import Path
from openpyxl import load_workbook
from .configExportExcel import exportar_con_plantilla

# Clase para importar y exportar datos
class ImportarExportarDatos():
  def __init__(self, carpeta, nombre_archivo):
    self.carpeta = Path(carpeta)
    self.nombre_archivo = nombre_archivo
    self.ruta_archivo = self.carpeta / self.nombre_archivo

  # Método que permite importar datos de objeto pickle
  def importar_pickle(self):
    with open(self.ruta_archivo, "rb") as archivo:  # Abre el archivo en modo binario
      base_datos_pickle = pickle.load(archivo)
    return base_datos_pickle

  # Método que permite importar datos de archivo excel
  def importar_excel(self,**kwargs):
    dataFrame = pd.read_excel(self.ruta_archivo, **kwargs)  # Usa ruta completa con pathlib
    return dataFrame

  # Método que permite exportar datos a objeto pickle
  def exportar_pickle(self, objeto):
    with open(self.ruta_archivo, "wb") as archivo:  # Usa el método `open` de `Path` para abrir el archivo en modo binario
      pickle.dump(objeto, archivo)

  # Método que permite exportar datos a archivo excel
  def exportar_excel(self, df, sheet_name, rewrite=False, ruta_formato_plantilla=None, **kwargs):
    """Exporta un DataFrame a una hoja específica de un archivo Excel."""
    self.carpeta.mkdir(parents=True, exist_ok=True)  # Asegurar que la carpeta existe

    # Si 'rewrite' es True y el archivo no existe, creamos uno vacío
    if rewrite and not self.ruta_archivo.exists():
      with pd.ExcelWriter(self.ruta_archivo, engine='openpyxl', mode='w') as writer:
        pd.DataFrame().to_excel(writer, sheet_name='EmptySheet') # Crear un archivo vacío con una hoja

    # Configurando las opciones de exportación
    if rewrite:
      options = {"mode": "a", "if_sheet_exists": "replace"}
    else:
      options = {"mode": "w"}

    # Exportando excel
    if ruta_formato_plantilla: # Si hay una plantilla definida, usar exportar_con_plantilla
      exportar_con_plantilla(df, ruta_formato_plantilla, self.ruta_archivo, sheet_name)
    else:
      with pd.ExcelWriter(self.ruta_archivo, engine='openpyxl', **options) as writer:
        df.to_excel(writer, sheet_name=sheet_name, **kwargs)

    # Eliminamos la hoja 'EmptySheet' creada inicialmente
    if rewrite:
      book = load_workbook(self.ruta_archivo)
      if 'EmptySheet' in book.sheetnames:
        del book['EmptySheet']
        book.save(self.ruta_archivo)

    # Aplicar configuraciones adicionales si están definidas
    if ruta_formato_plantilla:
      book = load_workbook(self.ruta_archivo)
      if col_fecha:
        config_fecha(book, col_fecha)
      if width_col:
        config_width_col(book, width_col)
      if align_col:
        config_align_col(book, align_col)
      book.save(self.ruta_archivo)



