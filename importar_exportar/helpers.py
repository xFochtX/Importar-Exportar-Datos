from pathlib import Path
from openpyxl import load_workbook
import zipfile
import xml.etree.ElementTree as ET

def ensure_folder(carpeta: Path):
  carpeta.mkdir(parents=True, exist_ok=True)

def delete_sheet(ruta_archivo, nombre_hoja):
  libro = load_workbook(ruta_archivo)
  if nombre_hoja in libro.sheetnames:
    del libro[nombre_hoja]
    libro.save(ruta_archivo)


def reorder_sheets(file_path, desired_order):
  """
  Reordena las hojas de un archivo Excel (.xlsx) según el orden especificado.
  Mantiene la estructura sin necesidad de abrir con openpyxl.
  """
  # Namespace principal del XML de Excel
  ns = {'main': 'http://schemas.openxmlformats.org/spreadsheetml/2006/main'}

  with zipfile.ZipFile(file_path, 'a') as z:
    # Extraer el contenido del workbook.xml
    workbook_xml = 'xl/workbook.xml'
    xml_content = z.read(workbook_xml)
    root = ET.fromstring(xml_content)

    # Buscar las hojas dentro del XML
    sheets = root.find('main:sheets', ns)
    if sheets is None:
      raise ValueError("No se encontró el nodo <sheets> en workbook.xml")

    # Convertir las hojas a lista editable
    sheet_elements = list(sheets)

    # Crear un diccionario: {nombre_hoja: elemento_xml}
    sheet_dict = {s.attrib['name']: s for s in sheet_elements}

    # Reconstruir las hojas según el orden deseado
    sheets[:] = [sheet_dict[name] for name in desired_order if name in sheet_dict]

    # Convertir el XML nuevamente a bytes
    new_xml_content = ET.tostring(root, encoding='utf-8', xml_declaration=True)

    # Sobrescribir workbook.xml dentro del ZIP (Excel)
    z.writestr(workbook_xml, new_xml_content)
