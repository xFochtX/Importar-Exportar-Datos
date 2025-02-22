from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment
from openpyxl.styles import NamedStyle, Font

# Método para colocar un formato específico a las fechas
def config_fecha(book, columna_fecha):
  # Establecer el formato de fecha en 'dd/mm/yyyy'
  date_style = NamedStyle(name='date_style')
  date_style.number_format = 'DD/MM/YYYY'
  date_style.font = Font(name='Calibri', size=11)  # Opcional: establecer la fuente y el tamaño de la fuente
  for sheetName in book.sheetnames:
    sheet = book[sheetName]
    # Obtener los nombres de las columnas desde la primera fila
    name_columns = [cell.value for cell in sheet[1]]
    pos_columna_fecha = name_columns.index(columna_fecha) + 1
    for cell in sheet[pos_columna_fecha]:
      cell.style = date_style

      # Método para configurar los anchos de las columnas


def config_width_col(book, width_col):
  for sheetName in book.sheetnames:
    sheet = book[sheetName]
    # Obtener los nombres de las columnas desde la primera fila
    name_columns = [cell.value for cell in sheet[1]]
    # Configurar el ancho de cada columna basado en los nombres de las columnas
    for idx, col_name in enumerate(name_columns, start=1):  # `start=1` para que el índice comience en 1
      if col_name in width_col:
        col_letter = sheet.cell(row=1, column=idx).column_letter  # Obtiene la letra identificadora de cada columna
        sheet.column_dimensions[col_letter].width = width_col[col_name]


# Método para configurar las alineaciones de las columnas
def config_align_col(book, align_col):
  for sheetName in book.sheetnames:
    sheet = book[sheetName]
    # Obtener los nombres de las columnas desde la primera fila
    name_columns = [cell.value for cell in sheet[1]]
    # Configurar la alineación de cada columna basado en los nombres de las columnas
    for idx, col_name in enumerate(name_columns, start=1):  # `start=1` para que el índice comience en 1
      if col_name in align_col:
        alignment = Alignment(horizontal=align_col[col_name], vertical='center')
        for cell in sheet[get_column_letter(idx)]:
          cell.alignment = alignment